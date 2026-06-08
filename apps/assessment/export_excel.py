# -*- coding: utf-8 -*-
"""
PoiExportExcel - Excel 成绩导出
对应 Java: PoiExportExcelController
5 种导出类型：
  1 = 实验统计（学校/课程/实验/班级/数量）
  2 = 实验成绩（学校/姓名/学科/实验/操作成绩）
  3 = 成绩统计（学科/实验/班级/优秀/中等/不合格人数）
  4 = 实验报告成绩（姓名/课程/资源名称/班级/上传次数/成绩）
  5 = 学生实验成绩（学号/班级/姓名/学科/实验/操作成绩/实验次数）
"""

import uuid
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import connection

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def _make_workbook(sheet_name: str, headers: list, rows: list) -> BytesIO:
    """生成 Excel 文件并返回 BytesIO"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头样式
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    cell_alignment = Alignment(horizontal='center', vertical='center')
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@csrf_exempt
@require_http_methods(['GET'])
def export_score(request):
    """
    GET /api/v1/assessment/export/score/
    导出 Excel 成绩文件
    
    参数:
        excelType: 1-5 对应不同导出类型
        un: 用户名（type=2 时需要）
        query/searchText: 搜索条件（可选）
        startDate/endDate: 日期范围（可选）
    """
    excel_type = request.GET.get('excelType', '')
    un = request.GET.get('un', '')
    search_text = request.GET.get('searchText') or request.GET.get('query', '')
    start_date = request.GET.get('startDate', '')
    end_date = request.GET.get('endDate', '')

    now = datetime.now()
    ts = int(now.timestamp() * 1000)

    if excel_type == '1':
        return _export_type1(request, ts)
    elif excel_type == '2':
        return _export_type2(request, ts, un)
    elif excel_type == '3':
        return _export_type3(request, ts)
    elif excel_type == '4':
        return _export_type4(request, ts)
    else:
        return _export_type5(request, ts)


def _export_type1(request, ts: int):
    """
    类型1：实验统计
    表头: 学校, 课程, 实验, 班级, 数量
    """
    headers = ['学校', '课程', '实验', '班级', '数量']
    rows = []

    # 从 session 读取（如果前端有设置）或查数据库
    score_data = request.session.get('scoreResultMap', {})
    poi_list = score_data.get('poiList', [])

    if poi_list:
        for item in poi_list:
            rows.append([
                item.get('schoolStr', ''),
                item.get('curriculumStr', ''),
                item.get('experimentStr', ''),
                item.get('classStr', ''),
                item.get('recordNum', 0),
            ])
    else:
        # 兜底：从 tb_experiment_record 统计
        with connection.cursor() as cur:
            cur.execute("""
                SELECT 
                    COALESCE(u.school_name, '') as school,
                    COALESCE(c.curriculum_name, '') as curriculum,
                    COALESCE(e.title, '') as experiment,
                    COALESCE(ci.class_card, '') as class_name,
                    COUNT(*) as cnt
                FROM tb_experiment_record er
                LEFT JOIN tb_user u ON er.user_Id = u.id
                LEFT JOIN tb_experiment e ON er.experiment_Id = e.id
                LEFT JOIN tb_curriculum c ON e.parent_id = c.id
                LEFT JOIN tb_class_info ci ON u.class_id = ci.id
                GROUP BY school, curriculum, experiment, class_name
                ORDER BY cnt DESC
                LIMIT 500
            """)
            for r in cur.fetchall():
                rows.append(list(r))

    output = _make_workbook('统计明细', headers, rows)
    return _excel_response(output, f"实验统计{ts}.xls")


def _export_type2(request, ts: int, un: str):
    """
    类型2：实验成绩
    表头: 学校, 姓名, 学科, 实验, 实验操作成绩
    """
    headers = ['学校', '姓名', '学科', '实验', '实验操作成绩']
    rows = []

    score_data = request.session.get('StuCJResultMap', {})
    score_list = score_data.get('list', [])

    if score_list:
        for item in score_list:
            rows.append([
                item.get('schoolStr', ''),
                item.get('studentName', ''),
                item.get('curriculumStr', ''),
                item.get('experimentStr', ''),
                item.get('operationScore', ''),
            ])
    else:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT 
                    COALESCE(u.school_name, ''),
                    COALESCE(u.name, ''),
                    COALESCE(c.curriculum_name, ''),
                    COALESCE(e.title, ''),
                    COALESCE(s.operation_score, 0)
                FROM tb_experiment_score s
                LEFT JOIN tb_user u ON s.user_id = u.id
                LEFT JOIN tb_experiment e ON s.experiment_id = e.id
                LEFT JOIN tb_curriculum c ON s.curriculum_id = c.id
                ORDER BY u.school_name
                LIMIT 500
            """)
            for r in cur.fetchall():
                rows.append(list(r))

    filename = f"{un}实验成绩{ts}.xls"
    output = _make_workbook('成绩明细', headers, rows)
    return _excel_response(output, filename)


def _export_type3(request, ts: int):
    """
    类型3：成绩统计
    表头: 学科, 实验, 班级, 优秀人数, 中等人数, 不合格人数
    """
    headers = ['学科', '实验', '班级', '优秀人数', '中等人数', '不合格人数']
    rows = []

    score_data = request.session.get('teaGLResultMap', {})
    tj_list = score_data.get('tJList', [])

    if tj_list:
        for item in tj_list:
            rows.append([
                item.get('curriculumStr', ''),
                item.get('experimentStr', ''),
                item.get('classStr', ''),
                item.get('excellent', 0),
                item.get('qualified', 0),
                item.get('unQualified', 0),
            ])
    else:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT 
                    COALESCE(c.curriculum_name, ''),
                    COALESCE(e.title, ''),
                    COALESCE(ci.class_card, ''),
                    COALESCE(r.excellent, 0),
                    COALESCE(r.qualified, 0),
                    COALESCE(r.unQualified, 0)
                FROM tb_experiment_score s
                LEFT JOIN tb_record_info r ON s.experiment_id = r.experiment_id
                LEFT JOIN tb_experiment e ON s.experiment_id = e.id
                LEFT JOIN tb_curriculum c ON s.curriculum_id = c.id
                LEFT JOIN tb_class_info ci ON s.class_Id = ci.id
                GROUP BY c.curriculum_name, e.title, ci.class_card, r.excellent, r.qualified, r.unQualified
                LIMIT 500
            """)
            for r in cur.fetchall():
                rows.append(list(r))

    output = _make_workbook('统计明细', headers, rows)
    return _excel_response(output, f"成绩统计{ts}.xls")


def _export_type4(request, ts: int):
    """
    类型4：实验报告成绩
    表头: 姓名, 课程, 资源名称, 班级, 上传次数, 成绩
    """
    headers = ['姓名', '课程', '资源名称', '班级', '上传次数', '成绩']
    rows = []

    score_data = request.session.get('bgscoreResultMap', {})
    bg_list = score_data.get('bglist', [])

    if bg_list:
        for item in bg_list:
            rows.append([
                item.get('un', ''),
                item.get('curriculumStr', ''),
                item.get('experimentStr', ''),
                item.get('classStr', ''),
                item.get('uploadNum', 0),
                item.get('reportScore', ''),
            ])
    else:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT 
                    COALESCE(r.un, ''),
                    COALESCE(c.curriculum_name, ''),
                    COALESCE(e.title, ''),
                    COALESCE(ci.class_card, ''),
                    COALESCE(r.upload_num, 0),
                    COALESCE(r.report_score, 0)
                FROM tb_experiment_report r
                LEFT JOIN tb_experiment e ON r.experiment_id = e.id
                LEFT JOIN tb_curriculum c ON r.curriculum_id = c.id
                LEFT JOIN tb_class_info ci ON r.class_Id = ci.id
                ORDER BY r.create_time DESC
                LIMIT 500
            """)
            for r in cur.fetchall():
                rows.append(list(r))

    output = _make_workbook('统计明细', headers, rows)
    return _excel_response(output, f"实验报告成绩{ts}.xls")


def _export_type5(request, ts: int):
    """
    类型5：学生实验成绩（默认）
    表头: 学号, 班级, 姓名, 学科, 实验, 实验操作成绩, 实验次数
    """
    headers = ['学号', '班级', '姓名', '学科', '实验', '实验操作成绩', '实验次数']
    rows = []

    score_data = request.session.get('teaCJResultMap', {})
    score_list = score_data.get('list', [])

    if score_list:
        for item in score_list:
            rows.append([
                item.get('idCard', ''),
                item.get('classStr', ''),
                item.get('un', ''),
                item.get('curriculumStr', ''),
                item.get('experimentStr', ''),
                item.get('operationScore', ''),
                item.get('experimentNum', ''),
            ])
    else:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT 
                    COALESCE(s.id_card, ''),
                    COALESCE(ci.class_card, ''),
                    COALESCE(u.name, ''),
                    COALESCE(c.curriculum_name, ''),
                    COALESCE(e.title, ''),
                    COALESCE(s.operation_score, 0),
                    COALESCE(s.experiment_num, 0)
                FROM tb_experiment_score s
                LEFT JOIN tb_user u ON s.user_id = u.id
                LEFT JOIN tb_experiment e ON s.experiment_id = e.id
                LEFT JOIN tb_curriculum c ON s.curriculum_id = c.id
                LEFT JOIN tb_class_info ci ON s.class_Id = ci.id
                ORDER BY u.name
                LIMIT 500
            """)
            for r in cur.fetchall():
                rows.append(list(r))

    output = _make_workbook('成绩明细', headers, rows)
    return _excel_response(output, f"学生实验成绩{ts}.xls")


def _excel_response(output: BytesIO, filename: str) -> HttpResponse:
    """生成 HTTP 响应，返回 Excel 文件下载"""
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.ms-excel'
    )
    # 文件名编码处理
    safe_filename = filename.encode('utf-8') if isinstance(filename, str) else filename
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
    response['Pragma'] = 'no-cache'
    response['Cache-Control'] = 'no-cache'
    return response
